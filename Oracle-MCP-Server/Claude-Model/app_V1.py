"""
Oracle NLQ Studio v3 — The Real Thing
======================================
Architecture (matching exactly what Claude Desktop does):

  Browser  ──REST──▶  Flask Backend  ──stdio──▶  sql -mcp (SQLcl MCP Server)
                           │
                           └──Anthropic API (tool_use)──▶ Claude
                                Claude calls SQLcl tools autonomously:
                                  list-connections → connect → run-sql → run-sql → ...

Claude IS the query engine. It reads the schema itself, builds SQL, executes it.
No separate NLQ→SQL translation. Claude drives the agentic loop exactly like Claude Desktop.

Setup:
    pip install flask flask-cors "mcp[cli]" anthropic openpyxl reportlab

Run:
    ANTHROPIC_API_KEY=sk-ant-...  SQLCL_PATH=/opt/sqlcl/bin/sql  python app.py
"""

import asyncio
import csv
import io
import json
import logging
import os
import re
import threading
from contextlib import AsyncExitStack
from datetime import datetime

import anthropic
from flask import Flask, Response, jsonify, request, send_file, stream_with_context
from flask_cors import CORS
from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client

# ── optional report libs ──────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# ─────────────────────────────────────────────────────────────────────────────
#  Single background event-loop  (Flask is sync; MCP SDK is async)
# ─────────────────────────────────────────────────────────────────────────────
_loop = asyncio.new_event_loop()
threading.Thread(target=_loop.run_forever, daemon=True).start()


def run_async(coro, timeout=120):
    return asyncio.run_coroutine_threadsafe(coro, _loop).result(timeout=timeout)


# ─────────────────────────────────────────────────────────────────────────────
#  SQLcl MCP Manager  — wraps the official MCP Python SDK
#  Launches  sql -mcp  as a subprocess, same as Claude Desktop config:
#    { "command": "/path/to/sql", "args": ["-mcp"] }
# ─────────────────────────────────────────────────────────────────────────────
class SQLclMCPManager:
    def __init__(self):
        self._session: ClientSession | None = None
        self._stack: AsyncExitStack | None = None
        self._tools: list[dict] = []          # raw MCP tool descriptors
        self._sqlcl_path = "sql"
        self._running = False
        self._active_conn: str | None = None

    # ── start / stop ──────────────────────────────────────────────────────────

    def start(self, sqlcl_path: str) -> dict:
        self._sqlcl_path = sqlcl_path
        try:
            return run_async(self._start())
        except Exception as e:
            log.exception("MCP start failed")
            return {"success": False, "error": str(e)}

    def stop(self):
        if self._running:
            try:
                run_async(self._stop())
            except Exception:
                pass
        self._running = False
        self._session = None
        self._active_conn = None

    @property
    def is_running(self):
        return self._running and self._session is not None

    # ── async internals ───────────────────────────────────────────────────────

    async def _start(self) -> dict:
        params = StdioServerParameters(
            command=self._sqlcl_path,   # e.g. /opt/sqlcl/bin/sql
            args=["-mcp"],              # MCP server mode
            env=dict(os.environ),       # forward JAVA_HOME etc.
        )
        stack = AsyncExitStack()
        read, write = await stack.enter_async_context(stdio_client(params))
        session: ClientSession = await stack.enter_async_context(ClientSession(read, write))

        # MCP handshake — mandatory
        init = await session.initialize()
        log.info("MCP ready: %s", init.serverInfo)

        # Discover tools (list-connections, connect, disconnect, run-sql, run-sqlcl)
        resp = await session.list_tools()
        self._tools = [
            {
                "name": t.name,
                "description": t.description or "",
                "input_schema": t.inputSchema if hasattr(t, "inputSchema") else {},
            }
            for t in resp.tools
        ]
        log.info("SQLcl tools: %s", [t["name"] for t in self._tools])

        self._stack = stack
        self._session = session
        self._running = True
        return {
            "success": True,
            "server": str(init.serverInfo),
            "tools": [t["name"] for t in self._tools],
        }

    async def _stop(self):
        if self._stack:
            await self._stack.aclose()
        self._stack = None
        self._session = None
        self._running = False

    async def _call(self, tool_name: str, arguments: dict) -> str:
        """Call one SQLcl MCP tool, return text result."""
        if not self._session:
            raise RuntimeError("MCP not running")
        result = await self._session.call_tool(tool_name, arguments=arguments)
        return "\n".join(
            block.text for block in result.content if hasattr(block, "text")
        )

    # ── public sync API ───────────────────────────────────────────────────────

    def call_tool(self, name: str, arguments: dict) -> str:
        return run_async(self._call(name, arguments))

    def list_connections(self) -> list[str]:
        raw = self.call_tool("list-connections", {})
        return _parse_connections(raw)

    def anthropic_tools(self) -> list[dict]:
        """
        Return SQLcl tools formatted for Anthropic tool_use API.
        Claude will decide when and how to call each one.
        """
        return [
            {
                "name": t["name"],
                "description": t["description"],
                "input_schema": t["input_schema"] or {
                    "type": "object",
                    "properties": {},
                },
            }
            for t in self._tools
        ]

    def get_active_conn(self): return self._active_conn
    def set_active_conn(self, v): self._active_conn = v
    def get_tools(self): return [t["name"] for t in self._tools]


mcp = SQLclMCPManager()


# ─────────────────────────────────────────────────────────────────────────────
#  The Agentic Loop
#  This is exactly what Claude Desktop does internally:
#   1. Send user message + SQLcl tools to Claude API
#   2. Claude responds with tool_use blocks (e.g. connect, run-sql)
#   3. We execute those tools against sql -mcp subprocess
#   4. Send tool results back to Claude
#   5. Repeat until Claude gives a final text answer
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an expert Oracle Database assistant with direct access to Oracle databases via SQLcl MCP tools.

## EFFICIENCY RULES — FOLLOW STRICTLY:
- **Minimise API roundtrips.** Plan all steps mentally first, then execute them with as few tool calls as possible.
- **Use schema-information FIRST** (not run-sql on USER_TABLES). It returns schema context in one call — never query USER_TABLES/USER_TAB_COLUMNS/ALL_TABLES manually unless schema-information is unavailable.
- **Do NOT loop through multiple exploratory queries.** If schema-information gives you enough context, go straight to the final query.
- **Maximum 4 tool calls per user message.** If you need more, explain why and ask the user to narrow the question.
- **Batch your thinking.** Decide the full SQL in one step after seeing schema, then execute once.

## WORKFLOW (follow this order):
1. If not connected → use list-connections ONCE, then connect ONCE.
2. If connected but schema unknown → call schema-information ONCE with the relevant table/owner hint.
3. Execute the final run-sql query ONCE.
4. Return the result with a clear explanation.

## DML SAFETY:
- For INSERT/UPDATE/DELETE/MERGE/DROP/TRUNCATE/CREATE/ALTER: show SQL first, state what it does, wait for user confirmation before executing.

## SQL STYLE:
- Always add /* Oracle NLQ Studio */ in SELECT queries.
- Use Oracle syntax: ROWNUM, NVL, TO_DATE, SYSDATE, CONNECT BY, DUAL.
- For Oracle ERP: AP_INVOICES_ALL, AP_SUPPLIERS, GL_JE_HEADERS, PER_ALL_PEOPLE_F etc.
- Qualify with schema owner when needed (e.g. ap.ap_invoices_all).

## OUTPUT:
- Present data as a clean markdown table.
- Include row count and any important observations.
- Keep explanations concise."""


def run_agent_turn(messages: list[dict], stream_callback=None) -> tuple[str, list[dict], list[dict]]:
    """
    Run one full agentic turn.
    Hard cap of MAX_ITERATIONS to prevent runaway 429s.
    Claude calls SQLcl tools autonomously but is instructed to be efficient.
    """
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    tools = mcp.anthropic_tools()
    tool_calls_log = []
    MAX_ITERATIONS = 6   # hard cap — prevents 429 storms
    iteration = 0

    while iteration < MAX_ITERATIONS:
        iteration += 1
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            tools=tools,
            messages=messages,
        )

        # Collect assistant content blocks
        assistant_content = []
        for block in response.content:
            if block.type == "text":
                assistant_content.append({"type": "text", "text": block.text})
            elif block.type == "tool_use":
                assistant_content.append({
                    "type": "tool_use",
                    "id": block.id,
                    "name": block.name,
                    "input": block.input,
                })

        messages.append({"role": "assistant", "content": assistant_content})

        # If Claude is done — no more tool calls
        if response.stop_reason == "end_turn":
            final_text = "\n".join(
                b["text"] for b in assistant_content if b.get("type") == "text"
            )
            return final_text, messages, tool_calls_log

        # Claude wants to call tools — execute each one against SQLcl MCP
        tool_results = []
        for block in assistant_content:
            if block.get("type") != "tool_use":
                continue

            tool_name = block["name"]
            tool_input = block["input"]
            tool_id = block["id"]

            log.info("Claude calling tool: %s(%s)", tool_name, tool_input)

            # ── DML safety: detect and flag ───────────────────────────────
            if tool_name == "run-sql":
                sql = tool_input.get("sql", "")
                is_dml = bool(re.match(
                    r"^\s*(INSERT|UPDATE|DELETE|MERGE|DROP|TRUNCATE|CREATE|ALTER)\b",
                    sql, re.I
                ))
                tool_input["_is_dml"] = is_dml  # flag for UI (stripped before call)

            # Execute the tool via the real MCP session
            try:
                clean_input = {k: v for k, v in tool_input.items() if not k.startswith("_")}
                result_text = mcp.call_tool(tool_name, clean_input)

                # Track active connection
                if tool_name == "connect":
                    mcp.set_active_conn(tool_input.get("connectionName", ""))

                tool_calls_log.append({
                    "tool": tool_name,
                    "input": tool_input,
                    "result_preview": result_text[:500],
                    "is_dml": tool_input.get("_is_dml", False),
                })

                if stream_callback:
                    stream_callback({"type": "tool_call", "tool": tool_name,
                                    "input": clean_input, "result": result_text[:300]})

            except Exception as e:
                result_text = f"Error calling {tool_name}: {e}"
                log.error(result_text)

            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tool_id,
                "content": result_text,
            })

        # Feed tool results back to Claude for next iteration
        messages.append({"role": "user", "content": tool_results})

    # Hit iteration cap
    log.warning("Agent hit MAX_ITERATIONS cap (%d calls)", MAX_ITERATIONS)
    partial = "\n".join(b["text"] for b in assistant_content if b.get("type") == "text") if assistant_content else ""
    if not partial:
        partial = ("⚠️ Query required too many steps. Be more specific — mention exact table/schema. "
                   f"Tools used: {[t["tool"] for t in tool_calls_log]}")
    return partial, messages, tool_calls_log



# ─────────────────────────────────────────────────────────────────────────────
#  Conversation Sessions  (in-memory, keyed by session_id)
# ─────────────────────────────────────────────────────────────────────────────
_sessions: dict[str, list[dict]] = {}
_last_result: dict = {"columns": [], "rows": []}
_query_history: list[dict] = []


def get_session(sid: str) -> list[dict]:
    if sid not in _sessions:
        _sessions[sid] = []
    return _sessions[sid]


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _parse_connections(raw: str) -> list[str]:
    names = []
    for line in raw.splitlines():
        c = re.sub(r"^[│├└─\s]+", "", line).strip()
        c = re.sub(r"\s*[\U0001F300-\U0001FFFF\U00002600-\U000027FF]\s*$", "", c).strip()
        if c and not re.match(r"(?i)^(connection|name|saved|#|\-)", c):
            names.append(c)
    return names


def _parse_sql_result(raw: str) -> dict:
    lines = [l for l in raw.splitlines() if l.strip()]
    sep = next((i for i, l in enumerate(lines) if re.match(r"^[-\s]+$", l) and i > 0), None)
    if sep:
        sep_line = lines[sep]
        spans, in_c, s = [], False, 0
        for i, ch in enumerate(sep_line):
            if ch == "-" and not in_c: s, in_c = i, True
            elif ch == " " and in_c: spans.append((s, i)); in_c = False
        if in_c: spans.append((s, len(sep_line)))
        def split_fw(ln): return [ln[a:b].strip() if a < len(ln) else "" for a, b in spans]
        cols = split_fw(lines[sep - 1])
        rows = []
        for ln in lines[sep + 1:]:
            if re.match(r"^\d+ row", ln, re.I): break
            rows.append(split_fw(ln))
        return {"columns": cols, "rows": rows, "raw": raw}
    return {"columns": [], "rows": [], "raw": raw, "message": raw}


# ─────────────────────────────────────────────────────────────────────────────
#  Report generators
# ─────────────────────────────────────────────────────────────────────────────

def gen_csv(cols, rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(cols); w.writerows(rows)
    return io.BytesIO(buf.getvalue().encode())

def gen_json(cols, rows):
    return io.BytesIO(json.dumps([dict(zip(cols, r)) for r in rows], indent=2, default=str).encode())

def gen_excel(cols, rows):
    if not OPENPYXL_OK: raise RuntimeError("pip install openpyxl")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Results"
    hf = PatternFill("solid", fgColor="C1440E")
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1): ws.cell(ri, ci, v)
    for cc in ws.columns:
        ws.column_dimensions[cc[0].column_letter].width = min(max(len(str(c.value or "")) for c in cc) + 4, 60)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def gen_pdf(cols, rows, title="Report"):
    if not REPORTLAB_OK: raise RuntimeError("pip install reportlab")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    tbl = Table([cols] + [[str(v) for v in r] for r in rows], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#C1440E")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#FFF5F2")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.grey),
    ]))
    doc.build([Paragraph(title, styles["Title"]), Spacer(1,10),
               Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), styles["Normal"]),
               Spacer(1,14), tbl])
    buf.seek(0); return buf


# ─────────────────────────────────────────────────────────────────────────────
#  Flask Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/health")
def health():
    return jsonify({
        "mcp_running": mcp.is_running,
        "active_connection": mcp.get_active_conn(),
        "tools": mcp.get_tools(),
        "anthropic_key": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "reportlab": REPORTLAB_OK,
        "openpyxl": OPENPYXL_OK,
    })


# ── MCP ───────────────────────────────────────────────────────────────────────

@app.route("/api/mcp/start", methods=["POST"])
def mcp_start():
    data = request.json or {}
    path = data.get("sqlcl_path") or os.environ.get("SQLCL_PATH", "sql")
    if mcp.is_running:
        return jsonify({"success": True, "message": "Already running", "tools": mcp.get_tools()})
    r = mcp.start(path)
    return jsonify(r), (200 if r.get("success") else 500)


@app.route("/api/mcp/stop", methods=["POST"])
def mcp_stop():
    mcp.stop()
    return jsonify({"success": True})


@app.route("/api/mcp/status")
def mcp_status():
    return jsonify({
        "running": mcp.is_running,
        "active_connection": mcp.get_active_conn(),
        "tools": mcp.get_tools(),
    })


# ── Connections ───────────────────────────────────────────────────────────────

@app.route("/api/connections")
def list_connections():
    if not mcp.is_running:
        return jsonify({"error": "MCP not running"}), 400
    try:
        return jsonify({"connections": mcp.list_connections(), "active": mcp.get_active_conn()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/connections/connect", methods=["POST"])
def connect_db():
    data = request.json or {}
    name = (data.get("connection_name") or "").strip()
    if not name: return jsonify({"error": "connection_name required"}), 400
    if not mcp.is_running: return jsonify({"error": "MCP not running"}), 400
    try:
        msg = mcp.call_tool("connect", {"connectionName": name})
        mcp.set_active_conn(name)
        return jsonify({"success": True, "message": msg, "active_connection": name})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Main Chat  (agentic loop) ─────────────────────────────────────────────────

@app.route("/api/chat", methods=["POST"])
def chat():
    """
    Core endpoint. User sends a message.
    Claude autonomously calls SQLcl tools (connect, run-sql, etc.)
    in a loop until it has a complete answer — exactly like Claude Desktop.
    """
    global _last_result

    data = request.json or {}
    session_id = data.get("session_id", "default")
    user_msg   = (data.get("message") or "").strip()
    confirmed  = data.get("dml_confirmed", False)  # DML confirmation flag

    if not user_msg:
        return jsonify({"error": "message required"}), 400
    if not mcp.is_running:
        return jsonify({"error": "Start MCP first"}), 400
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"error": "ANTHROPIC_API_KEY not set"}), 400

    messages = get_session(session_id)

    # Append DML confirmation note if needed
    if confirmed:
        user_msg = user_msg + "\n\n[User has confirmed: Please proceed and execute the DML operation.]"

    messages.append({"role": "user", "content": user_msg})

    try:
        final_text, messages, tool_calls = run_agent_turn(messages)

        # Update session
        _sessions[session_id] = messages

        # Try to extract tabular result from last run-sql call
        sql_result = None
        last_sql = None
        has_dml = False
        for tc in reversed(tool_calls):
            if tc["tool"] == "run-sql":
                if tc.get("is_dml"):
                    has_dml = True
                raw = tc.get("result_preview", "")
                parsed = _parse_sql_result(raw)
                if parsed.get("columns"):
                    sql_result = parsed
                    _last_result = parsed
                last_sql = tc["input"].get("sql", "")
                break

        # Record history
        _query_history.append({
            "timestamp": datetime.now().isoformat(),
            "user_message": user_msg[:120],
            "tool_calls": len(tool_calls),
            "sql": last_sql or "",
        })
        if len(_query_history) > 100:
            _query_history.pop(0)

        return jsonify({
            "response": final_text,
            "tool_calls": tool_calls,
            "sql_result": sql_result,
            "has_dml": has_dml,
            "active_connection": mcp.get_active_conn(),
        })

    except Exception as e:
        log.exception("chat failed")
        # Remove the failed user message so session stays clean
        if messages and messages[-1]["role"] == "user":
            messages.pop()
        return jsonify({"error": str(e)}), 500


@app.route("/api/chat/clear", methods=["POST"])
def clear_chat():
    data = request.json or {}
    sid = data.get("session_id", "default")
    _sessions[sid] = []
    return jsonify({"success": True})


@app.route("/api/history")
def history():
    return jsonify({"history": list(reversed(_query_history))})


# ── Reports ───────────────────────────────────────────────────────────────────

@app.route("/api/report/<fmt>", methods=["POST"])
def export_report(fmt):
    data  = request.json or {}
    cols  = data.get("columns") or _last_result.get("columns", [])
    rows  = data.get("rows")    or _last_result.get("rows", [])
    title = data.get("title", "Oracle NLQ Studio")
    if not cols: return jsonify({"error": "No data — ask a query first"}), 400
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        match fmt.lower():
            case "csv":  return send_file(gen_csv(cols, rows),   mimetype="text/csv", as_attachment=True, download_name=f"report_{ts}.csv")
            case "json": return send_file(gen_json(cols, rows),  mimetype="application/json", as_attachment=True, download_name=f"report_{ts}.json")
            case "xlsx": return send_file(gen_excel(cols, rows), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"report_{ts}.xlsx")
            case "pdf":  return send_file(gen_pdf(cols, rows, title), mimetype="application/pdf", as_attachment=True, download_name=f"report_{ts}.pdf")
            case _: return jsonify({"error": f"Unknown format {fmt}"}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    log.info("Oracle NLQ Studio v3 on :%d", port)
    log.info("SQLcl path: %s", os.environ.get("SQLCL_PATH", "sql (from PATH)"))
    log.info("Anthropic key: %s", "SET" if os.environ.get("ANTHROPIC_API_KEY") else "NOT SET")
    app.run(host="0.0.0.0", port=port, debug=False)
